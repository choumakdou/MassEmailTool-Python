"""
Mass Email Marketing Tool (v6)
------------------------------
Same modern dark-mode desktop UI as v5, with one addition: a
"Custom SMTP" provider option that lets the user point at any
SMTP server (custom domain on Microsoft 365, Google Workspace,
Zoho, self-hosted Exchange, etc.).

SSL/TLS is auto-detected by port:
  - 465 -> SMTP_SSL  (implicit TLS)
  - all others -> STARTTLS (opportunistic TLS)
"""
import os
import sys
import json
import smtplib
import ssl
import queue
import subprocess
import threading
from datetime import datetime

import tkinter as tk
from tkinter import messagebox

import customtkinter as ctk

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from email.mime.base import MIMEBase
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


# ============================ Config ============================

SETTINGS_FILE   = "email_settings.json"
TEMPLATE_FILE   = "Email_List.xlsx"
ATTACHMENTS_DIR = "Attachments"

# Each provider is a label + (server, port). The "custom" entry has an
# empty server — filled in by the user at runtime. Port 465 is treated
# as implicit SSL automatically; everything else uses STARTTLS.
PROVIDERS = {
    "gmail":     {"label": "Gmail",         "server": "smtp.gmail.com",         "port": 587},
    "outlook":   {"label": "Outlook.com",   "server": "smtp-mail.outlook.com",  "port": 587},
    "office365": {"label": "Office 365",    "server": "smtp.office365.com",     "port": 587},
    "custom":    {"label": "Custom SMTP",   "server": "",                       "port": 587},
}

COL_NAME, COL_EMAIL, COL_SUBJECT, COL_BODY = 1, 2, 3, 4
COL_ATT1, COL_ATT2, COL_STATUS            = 5, 6, 7


# ============================ Helpers ============================

def get_exe_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_settings_path()   : return os.path.join(get_exe_dir(), SETTINGS_FILE)
def get_template_path()   : return os.path.join(get_exe_dir(), TEMPLATE_FILE)
def get_attachments_dir() : return os.path.join(get_exe_dir(), ATTACHMENTS_DIR)


def load_settings():
    p = get_settings_path()
    if not os.path.exists(p):
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        if all(k in data for k in ("provider", "email", "password")):
            return data
    except Exception:
        pass
    return None


def save_settings(data):
    with open(get_settings_path(), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def create_template() -> str:
    path = get_template_path()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Email List"
    headers = ["Name", "Email", "Subject", "Message",
               "Attachment 1", "Attachment 2", "Status"]
    for col, h in enumerate(headers, 1):
        c = sheet.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")
    example = [
        "(example — delete this row)",
        "you@example.com",
        "Sample subject",
        "Hi there,\n\nThis is what a message looks like.",
        "flyer.pdf",
        "",
        "Sent",
    ]
    for col, val in enumerate(example, 1):
        sheet.cell(row=2, column=col, value=val)
    for col_letter, width in zip("ABCDEFG", [18, 30, 30, 50, 25, 25, 12]):
        sheet.column_dimensions[col_letter].width = width
    wb.save(path)
    return path


def resolve_attachment(name):
    if not name:
        return None
    s = str(name).strip()
    if not s:
        return None
    candidate = os.path.join(get_attachments_dir(), os.path.basename(s))
    if os.path.exists(candidate):
        return candidate
    if os.path.exists(s):
        return s
    return None


def attach_file(msg, filepath):
    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{os.path.basename(filepath)}"')
    msg.attach(part)


def get_smtp_target(settings):
    """Return (server, port, use_ssl) for the current settings.

    - For named providers, look up in PROVIDERS.
    - For "custom", pull server/port from the saved settings.
    - Port 465 -> implicit SSL; everything else -> STARTTLS.
    """
    if settings.get("provider") == "custom":
        server = (settings.get("server") or "").strip()
        port   = int(settings.get("port") or 587)
    else:
        cfg    = PROVIDERS[settings["provider"]]
        server = cfg["server"]
        port   = cfg["port"]
    return server, port, (port == 465)


def smtp_handshake(settings, timeout=30):
    server_addr, port, use_ssl = get_smtp_target(settings)
    ctx = ssl.create_default_context()
    if use_ssl:
        server = smtplib.SMTP_SSL(server_addr, port, timeout=timeout, context=ctx)
    else:
        server = smtplib.SMTP(server_addr, port, timeout=timeout)
    server.ehlo()
    if not use_ssl:
        server.starttls(context=ctx)
        server.ehlo()
    server.login(settings["email"], settings["password"])
    return server


def open_in_default_app(filepath):
    if not os.path.exists(filepath):
        return False
    try:
        if sys.platform == "win32":
            os.startfile(str(filepath))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(filepath)])
        else:
            subprocess.Popen(["xdg-open", str(filepath)])
        return True
    except Exception:
        return False


def count_unsent_rows():
    tpath = get_template_path()
    if not os.path.exists(tpath):
        return 0
    try:
        wb = openpyxl.load_workbook(tpath, read_only=True, data_only=True)
        sheet = wb.active
        n = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[1] and str(row[6] or "").strip().lower() != "sent":
                n += 1
        wb.close()
        return n
    except Exception:
        return 0


# ============================ GUI ============================

class EmailToolApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.title("Mass Email Marketing Tool")
        self.geometry("860x760")
        self.minsize(740, 680)

        self.log_queue: "queue.Queue[str]" = queue.Queue()
        self.running = False
        self.stop_flag = False
        self.settings = None
        self.total_to_send = 0
        self.done_count = 0
        self.error_count = 0

        self._build_ui()
        self._bootstrap()
        self._poll_log_queue()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---------- UI construction ----------
    def _build_ui(self):
        # Header
        ctk.CTkLabel(self, text="📧 Mass Email Marketing Tool",
                     font=ctk.CTkFont(size=22, weight="bold")
                     ).pack(pady=(18, 2), padx=20, anchor="w")
        ctk.CTkLabel(self,
                     text="Personalized bulk email from a spreadsheet · Gmail / Outlook / Office 365 / Custom SMTP",
                     text_color=("gray60", "gray70"),
                     font=ctk.CTkFont(size=12)
                     ).pack(padx=22, anchor="w")

        # ---- Credentials ----
        creds = ctk.CTkFrame(self)
        creds.pack(fill="x", padx=16, pady=(14, 6))
        ctk.CTkLabel(creds, text="Sending Account",
                     font=ctk.CTkFont(size=14, weight="bold")
                     ).pack(anchor="w", padx=14, pady=(10, 4))

        # Provider
        row1 = ctk.CTkFrame(creds, fg_color="transparent")
        row1.pack(fill="x", padx=14, pady=4)
        ctk.CTkLabel(row1, text="Provider", width=90, anchor="w").pack(side="left")
        self.provider_var = tk.StringVar(value="office365")
        ctk.CTkOptionMenu(row1, variable=self.provider_var,
                          values=list(PROVIDERS.keys()),
                          command=self._on_provider_change,
                          width=240).pack(side="left")

        # Custom SMTP fields (hidden unless "custom" is selected)
        self.custom_frame = ctk.CTkFrame(creds, fg_color="transparent")
        crow = ctk.CTkFrame(self.custom_frame, fg_color="transparent")
        crow.pack(fill="x", pady=2)
        ctk.CTkLabel(crow, text="SMTP Server", width=90, anchor="w",
                     text_color=("gray40", "gray70")
                     ).pack(side="left")
        self.custom_server_var = tk.StringVar()
        ctk.CTkEntry(crow, textvariable=self.custom_server_var, width=260,
                     placeholder_text="smtp.your-company.com"
                     ).pack(side="left", padx=(0, 12))
        ctk.CTkLabel(crow, text="Port", anchor="w",
                     text_color=("gray40", "gray70")
                     ).pack(side="left")
        self.custom_port_var = tk.StringVar(value="587")
        ctk.CTkEntry(crow, textvariable=self.custom_port_var, width=70
                     ).pack(side="left", padx=(4, 0))
        ctk.CTkLabel(self.custom_frame,
                     text="Tip: 465 = implicit SSL · 587/25/2525 = STARTTLS",
                     text_color=("gray50", "gray60"),
                     font=ctk.CTkFont(size=11)
                     ).pack(anchor="w", padx=104, pady=(0, 4))

        # Email
        row2 = ctk.CTkFrame(creds, fg_color="transparent")
        row2.pack(fill="x", padx=14, pady=4)
        ctk.CTkLabel(row2, text="Email", width=90, anchor="w").pack(side="left")
        self.email_var = tk.StringVar()
        ctk.CTkEntry(row2, textvariable=self.email_var, width=340,
                     placeholder_text="you@your-company.com").pack(side="left")

        # Password
        row3 = ctk.CTkFrame(creds, fg_color="transparent")
        row3.pack(fill="x", padx=14, pady=4)
        ctk.CTkLabel(row3, text="App Password", width=90, anchor="w").pack(side="left")
        self.password_var = tk.StringVar()
        self.password_entry = ctk.CTkEntry(row3, textvariable=self.password_var,
                                            width=340, show="*",
                                            placeholder_text="account password (or App Password if MFA is on)")
        self.password_entry.pack(side="left")
        self.show_pw = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(row3, text="Show", variable=self.show_pw,
                        width=60, command=self._toggle_pw).pack(side="left", padx=(8, 0))

        # Buttons row
        btnrow = ctk.CTkFrame(creds, fg_color="transparent")
        btnrow.pack(fill="x", padx=14, pady=(10, 12))
        ctk.CTkButton(btnrow, text="💾  Save",  command=self._save_settings,
                      width=110).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btnrow, text="🔌  Test Connection",
                      command=self._test_connection, width=160,
                      fg_color="transparent", border_width=1,
                      text_color=("gray10", "gray90")
                      ).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btnrow, text="🗑  Clear Saved",
                      command=self._clear_settings, width=130,
                      fg_color="transparent", border_width=1,
                      text_color=("gray10", "gray90")
                      ).pack(side="left")

        # ---- Recipients & files ----
        files = ctk.CTkFrame(self)
        files.pack(fill="x", padx=16, pady=6)
        ctk.CTkLabel(files, text="Recipients & Attachments",
                     font=ctk.CTkFont(size=14, weight="bold")
                     ).pack(anchor="w", padx=14, pady=(10, 4))
        self.files_status = ctk.CTkLabel(files, text="", justify="left",
                                          font=ctk.CTkFont(size=12))
        self.files_status.pack(anchor="w", padx=14, pady=4)

        btnrow2 = ctk.CTkFrame(files, fg_color="transparent")
        btnrow2.pack(fill="x", padx=14, pady=(4, 12))
        ctk.CTkButton(btnrow2, text="📄  Open Spreadsheet",
                      command=self._open_template, width=170).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btnrow2, text="📎  Open Attachments Folder",
                      command=self._open_attachments, width=200).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btnrow2, text="🔄  Refresh",
                      command=self._refresh_file_status, width=110,
                      fg_color="transparent", border_width=1,
                      text_color=("gray10", "gray90")
                      ).pack(side="left")

        # ---- Activity log ----
        log_frame = ctk.CTkFrame(self)
        log_frame.pack(fill="both", expand=True, padx=16, pady=6)
        ctk.CTkLabel(log_frame, text="Activity Log",
                     font=ctk.CTkFont(size=14, weight="bold")
                     ).pack(anchor="w", padx=14, pady=(10, 4))
        self.log_textbox = ctk.CTkTextbox(log_frame, height=200,
                                          font=ctk.CTkFont(family="Consolas", size=11),
                                          wrap="word")
        self.log_textbox.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        self.log_textbox.configure(state="disabled")

        # ---- Progress ----
        prog_row = ctk.CTkFrame(self, fg_color="transparent")
        prog_row.pack(fill="x", padx=16, pady=(0, 4))
        self.progress_label = ctk.CTkLabel(prog_row, text="Idle", anchor="w",
                                            font=ctk.CTkFont(size=12))
        self.progress_label.pack(side="left")
        self.progress_bar = ctk.CTkProgressBar(prog_row)
        self.progress_bar.pack(side="right", fill="x", expand=True, padx=(12, 0))
        self.progress_bar.set(0)

        # ---- Action bar ----
        action = ctk.CTkFrame(self, fg_color="transparent")
        action.pack(fill="x", padx=16, pady=(4, 16))
        self.run_btn = ctk.CTkButton(action, text="▶  Run",
                                      command=self._on_run, height=42,
                                      font=ctk.CTkFont(size=14, weight="bold"))
        self.run_btn.pack(side="left", padx=(0, 8), fill="x", expand=True)
        self.stop_btn = ctk.CTkButton(action, text="■  Stop",
                                       command=self._on_stop, height=42,
                                       fg_color="#B33A3A", hover_color="#8B2A2A",
                                       state="disabled",
                                       font=ctk.CTkFont(size=14, weight="bold"))
        self.stop_btn.pack(side="left", padx=(0, 8), fill="x", expand=True)
        ctk.CTkButton(action, text="Quit", command=self._on_close, height=42,
                       width=100, fg_color="transparent", border_width=1,
                       text_color=("gray10", "gray90")
                       ).pack(side="left")

    # ---------- Provider change ----------
    def _on_provider_change(self, value):
        if value == "custom":
            self.custom_frame.pack(fill="x", padx=14, pady=4)
        else:
            self.custom_frame.pack_forget()

    # ---------- Bootstrap ----------
    def _bootstrap(self):
        os.makedirs(get_attachments_dir(), exist_ok=True)
        if not os.path.exists(get_template_path()):
            create_template()
            self._log(f"[setup] Created template → {get_template_path()}")
            self._log(f"[setup] Created attachments folder → {get_attachments_dir()}")
        self.settings = load_settings()
        if self.settings:
            self.provider_var.set(self.settings.get("provider", "office365"))
            self.email_var.set(self.settings.get("email", ""))
            self.password_var.set(self.settings.get("password", ""))
            if self.settings.get("provider") == "custom":
                self.custom_server_var.set(self.settings.get("server", ""))
                self.custom_port_var.set(str(self.settings.get("port", 587)))
            self._log(f"[setup] Loaded saved credentials for {self.settings.get('email')}")
        else:
            self._log("[setup] No saved credentials yet — fill in the form above.")
        # Make sure the custom frame is shown iff needed.
        self._on_provider_change(self.provider_var.get())
        self._refresh_file_status()

    def _refresh_file_status(self):
        tpath = get_template_path()
        adir  = get_attachments_dir()
        template_ok = os.path.exists(tpath)
        att_count = 0
        if os.path.isdir(adir):
            att_count = sum(1 for f in os.listdir(adir)
                            if os.path.isfile(os.path.join(adir, f))
                            and not f.startswith("."))
        row_count = count_unsent_rows() if template_ok else 0
        self.total_to_send = row_count
        status = (f"📄  Spreadsheet: {os.path.basename(tpath)}"
                  f"   ·   ✉️  {row_count} recipient(s) ready"
                  f"\n📎  Attachments folder: {att_count} file(s)"
                  f"   ·   📁  {adir}")
        self.files_status.configure(text=status)

    # ---------- Logging (thread-safe) ----------
    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_queue.put(f"{ts}  {msg}")

    def _poll_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self.log_textbox.configure(state="normal")
                self.log_textbox.insert("end", msg + "\n")
                self.log_textbox.see("end")
                self.log_textbox.configure(state="disabled")
        except queue.Empty:
            pass
        self.after(60, self._poll_log_queue)

    # ---------- Credentials ----------
    def _toggle_pw(self):
        self.password_entry.configure(show="" if self.show_pw.get() else "*")

    def _collect_settings(self):
        provider = self.provider_var.get().strip().lower()
        email = self.email_var.get().strip()
        password = self.password_var.get()
        if provider not in PROVIDERS:
            return None
        if not email or not password:
            return None
        out = {"provider": provider, "email": email, "password": password}
        if provider == "custom":
            server = self.custom_server_var.get().strip()
            try:
                port = int(self.custom_port_var.get().strip() or "587")
            except ValueError:
                port = 587
            if not server:
                return None
            out["server"] = server
            out["port"]   = port
        return out

    def _save_settings(self):
        s = self._collect_settings()
        if not s:
            messagebox.showwarning("Missing info",
                "Please fill in email + password.\n"
                "For Custom SMTP, also fill in the server and port.")
            return
        save_settings(s)
        self.settings = s
        self._log(f"[settings] Saved credentials for {s['email']} ({s['provider']})")
        messagebox.showinfo("Saved", f"Credentials saved to:\n{get_settings_path()}")

    def _clear_settings(self):
        p = get_settings_path()
        if os.path.exists(p):
            os.remove(p)
        self.settings = None
        self.password_var.set("")
        self._log("[settings] Cleared saved credentials")

    def _test_connection(self):
        s = self._collect_settings()
        if not s:
            messagebox.showwarning("Missing info",
                "Please fill in email + password.\n"
                "For Custom SMTP, also fill in the server and port.")
            return
        try:
            srv, port, ssl_used = get_smtp_target(s)
        except Exception as e:
            messagebox.showerror("Bad settings", str(e)); return
        self._log(f"[test] Connecting to {srv}:{port} "
                  f"({'SSL' if ssl_used else 'STARTTLS'}) as {s['email']} ...")
        def worker():
            try:
                server = smtp_handshake(s, timeout=15)
                server.quit()
                self._log("[test] ✅  Connection successful.")
                self.after(0, lambda: messagebox.showinfo("Connection OK",
                    f"Connected to {srv}:{port} as {s['email']}."))
            except Exception as e:
                self._log(f"[test] ❌  FAILED: {e}")
                self.after(0, lambda e=e: messagebox.showerror("Connection failed",
                    f"{e}\n\nTip: with MFA enabled, you usually need an App Password "
                    f"rather than your account password.\n"
                    f"Microsoft: https://mysignins.microsoft.com/security-info\n"
                    f"Google:    https://myaccount.google.com/apppasswords"))
        threading.Thread(target=worker, daemon=True).start()

    # ---------- Open files / folders ----------
    def _open_template(self):
        if not os.path.exists(get_template_path()):
            create_template()
        open_in_default_app(get_template_path())

    def _open_attachments(self):
        os.makedirs(get_attachments_dir(), exist_ok=True)
        open_in_default_app(get_attachments_dir())

    # ---------- Run / Stop ----------
    def _on_run(self):
        if self.running:
            return
        s = self._collect_settings()
        if not s:
            messagebox.showwarning("Missing info",
                "Please fill in email + password.\n"
                "For Custom SMTP, also fill in the server and port.")
            return
        self.settings = s
        save_settings(s)
        self._refresh_file_status()
        if self.total_to_send == 0:
            if not messagebox.askyesno("Nothing to send",
                    "No recipients with Status != 'Sent' were found.\n\nRun anyway?"):
                return
        else:
            srv, port, ssl_used = get_smtp_target(s)
            via = f" via {s['provider']}" if s['provider'] != "custom" else f" via {srv}:{port}"
            if not messagebox.askyesno("Confirm",
                    f"About to send {self.total_to_send} email(s){via} as {s['email']}.\n\nProceed?"):
                return
        self.stop_flag = False
        self.done_count = 0
        self.error_count = 0
        self.progress_bar.set(0)
        self.progress_label.configure(text="Starting…")
        self.run_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.running = True
        threading.Thread(target=self._send_worker, daemon=True).start()

    def _on_stop(self):
        if not self.running:
            return
        self.stop_flag = True
        self._log("[stop] Stop requested — will halt after the current email.")

    def _send_worker(self):
        try:
            tpath = get_template_path()
            wb = openpyxl.load_workbook(tpath)
            sheet = wb.active
            try:
                server = smtp_handshake(self.settings)
                srv, port, ssl_used = get_smtp_target(self.settings)
                mode = "SSL" if ssl_used else "STARTTLS"
                self._log(f"[send] Connected to {srv}:{port} ({mode}) as {self.settings['email']}.")
            except Exception as e:
                self._log(f"[fatal] SMTP connection failed: {e}")
                self.after(0, lambda: messagebox.showerror("SMTP failed",
                    f"Could not connect:\n{e}\n\nCheck credentials / App Password."))
                return
            self._log(f"[send] Sending to {self.total_to_send} recipient(s)…")
            try:
                for row in range(2, sheet.max_row + 1):
                    if self.stop_flag:
                        self._log("[stop] Halted by user.")
                        break
                    name    = sheet.cell(row=row, column=COL_NAME).value
                    email   = sheet.cell(row=row, column=COL_EMAIL).value
                    subject = sheet.cell(row=row, column=COL_SUBJECT).value
                    body    = sheet.cell(row=row, column=COL_BODY).value
                    att1    = sheet.cell(row=row, column=COL_ATT1).value
                    att2    = sheet.cell(row=row, column=COL_ATT2).value
                    status  = sheet.cell(row=row, column=COL_STATUS).value
                    if not email or str(status or "").strip().lower() == "sent":
                        continue
                    self._log(f"→  {email}  …")
                    try:
                        msg = MIMEMultipart()
                        msg["From"]    = self.settings["email"]
                        msg["To"]      = str(email)
                        msg["Subject"] = str(subject) if subject else "Marketing Update"
                        text = (f"Dear {name if name else 'Client'},\n\n"
                                f"{body if body else ''}\n\n"
                                "Best regards,\nMarketing Team")
                        msg.attach(MIMEText(text, "plain"))
                        for att in (att1, att2):
                            p = resolve_attachment(att)
                            if p:
                                attach_file(msg, p)
                            elif att and str(att).strip():
                                self._log(f"      (attachment missing: {att})")
                        server.sendmail(self.settings["email"], str(email), msg.as_string())
                        sheet.cell(row=row, column=COL_STATUS, value="Sent")
                        self.done_count += 1
                        self._log(f"   ✅  sent ({self.done_count} ok / {self.error_count} err)")
                    except Exception as e:
                        sheet.cell(row=row, column=COL_STATUS, value="Error")
                        self.error_count += 1
                        self._log(f"   ❌  failed: {e}")
                    total = self.done_count + self.error_count
                    progress = total / max(self.total_to_send, 1)
                    self.after(0, lambda p=progress: self.progress_bar.set(p))
                    self.after(0, lambda: self.progress_label.configure(
                        text=f"Sent {self.done_count}  ·  Failed {self.error_count}  ·  Total {self.total_to_send}"))
            finally:
                try: server.quit()
                except Exception: pass
                try:
                    wb.save(tpath)
                except Exception as e:
                    self._log(f"[warn] Could not save Excel: {e}")
            self._log(f"[done] ✅ Sent: {self.done_count}   ❌ Failed: {self.error_count}")
            self.after(0, lambda: messagebox.showinfo("Done",
                f"Sent: {self.done_count}\nFailed: {self.error_count}"))
        finally:
            self.running = False
            self.after(0, lambda: self.run_btn.configure(state="normal"))
            self.after(0, lambda: self.stop_btn.configure(state="disabled"))
            self.after(0, self._refresh_file_status)

    # ---------- Close ----------
    def _on_close(self):
        if self.running:
            if not messagebox.askyesno("Quit", "A send is in progress. Quit anyway?"):
                return
        self.destroy()


# ============================ Entry point ============================

def main():
    try:
        app = EmailToolApp()
        app.mainloop()
    except Exception as e:
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Startup failed",
                f"{e}\n\nIf this mentions 'customtkinter', make sure it is installed:\n"
                f"  pip install customtkinter")
        except Exception:
            print(f"FATAL: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
