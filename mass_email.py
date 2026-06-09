import openpyxl
import os
import sys
import tkinter as tk
from tkinter import filedialog

# Required for Outlook interaction on Windows
try:
    import win32com.client as win32
except ImportError:
    print("Error: 'pywin32' not found. This tool must run on Windows with Outlook.")
    input("Press Enter to exit...")
    sys.exit(1)

def get_excel_path():
    """Finds the Excel file: either by user selection or looking in the current folder."""
    # 1. Try to find any .xlsx file in the current folder automatically
    current_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    print(f"Checking for Excel files in: {current_dir}")
    
    excel_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    if len(excel_files) == 1:
        auto_file = os.path.join(current_dir, excel_files[0])
        print(f"Found exactly one Excel file: {excel_files[0]}")
        return auto_file
    
    # 2. If zero or multiple files, ask the user to pick one
    print("No unique Excel file found automatically. Please select your file...")
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True) # Bring to front
    file_path = filedialog.askopenfilename(
        title="Select your Marketing Excel List",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    root.destroy()
    return file_path

def run_tool():
    print("========================================")
    print("   MASS EMAIL MARKETING TOOL (v3)       ")
    print("========================================\n")
    
    excel_path = get_excel_path()
    
    if not excel_path:
        print("Error: No file selected. Exiting.")
        return

    print(f"\nTarget File: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        print(f"Successfully loaded sheet: {sheet.title}")
    except Exception as e:
        print(f"CRITICAL ERROR: Could not open Excel file.\nReason: {e}")
        return

    try:
        outlook = win32.Dispatch('outlook.application')
        print("Connected to Outlook successfully.")
    except Exception as e:
        print(f"CRITICAL ERROR: Could not connect to Outlook.\nEnsure Outlook is installed and open.\nReason: {e}")
        return

    sent_count = 0
    error_count = 0

    # Start from row 2 (skipping headers)
    # A=Name, B=Email, C=Subject, D=Update, E=Path1, F=Path2, G=Status
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        email = sheet.cell(row=row, column=2).value
        subject = sheet.cell(row=row, column=3).value
        update = sheet.cell(row=row, column=4).value
        path1 = sheet.cell(row=row, column=5).value
        path2 = sheet.cell(row=row, column=6).value
        status = sheet.cell(row=row, column=7).value

        if not email or str(status).strip().lower() == "sent":
            continue

        print(f"Processing: {email}...", end=" ")
        
        try:
            mail = outlook.CreateItem(0)
            mail.To = str(email)
            mail.Subject = str(subject) if subject else "Marketing Update"
            
            body = f"Dear {name if name else 'Client'},\n\n{update if update else ''}\n\nBest regards,\nMarketing Team"
            mail.Body = body

            # Attachments
            for path in [path1, path2]:
                if path:
                    p = str(path).strip()
                    if os.path.exists(p):
                        mail.Attachments.Add(os.path.abspath(p))
                    else:
                        print(f"(Warning: File not found at {p})", end=" ")

            # Change .Display() to .Send() once you are ready to go fully automatic
            mail.Display() 
            
            sheet.cell(row=row, column=7).value = "Sent"
            sent_count += 1
            print("OK")

        except Exception as e:
            print(f"FAILED: {e}")
            sheet.cell(row=row, column=7).value = "Error"
            error_count += 1

    try:
        wb.save(excel_path)
        print("\n" + "="*40)
        print(f"COMPLETED!")
        print(f"Emails Prepared: {sent_count}")
        print(f"Errors: {error_count}")
        print("Excel status updated.")
        print("="*40)
    except Exception as e:
        print(f"\nWarning: Could not save status to Excel. Is it open in another window?\nError: {e}")

if __name__ == "__main__":
    try:
        run_tool()
    except Exception as e:
        print(f"\nAn unexpected error occurred: {e}")
    
    print("\nPress Enter to close this window...")
    input()
