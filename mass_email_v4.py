"""
Mass Email Marketing Tool (v4)
------------------------------
On first run it bootstraps everything in the folder next to the .exe:
  • Email_List.xlsx   — input template (one row per recipient)
  • Attachments/      — folder where you drop files referenced by the template
  • email_settings.json — saved SMTP credentials (delete to reset)

It then connects via SMTP (Gmail / Outlook.com / Office 365) using the
email + password (or App Password) you provide, and sends every row whose
Status is not "Sent".

No local Outlook install required, no need to know the columns in advance.
"""

import os
import sys
import json
import smtplib
import ssl
import tkinter as tk
from tkinter import simpledialog

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from email.mime.base import MIMEBase
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


# ----------------------------- Configuration -----------------------------

SETTINGS_FILE    = "email_settings.json"
TEMPLATE_FILE    = "Email_List.xlsx"
ATTACHMENTS_DIR  = "Attachments"

# SMTP presets for the supported providers.
PROVIDERS = {
    "gmail":     {"server": "smtp.gmail.com",         "port": 587},
    "outlook":   {"server": "smtp-mail.outlook.com",  "port": 587},  # outlook.com / hotmail / live
    "office365": {"server": "smtp.office365.com",     "port": 587},  # Microsoft 365
}

# Excel column layout
COL_NAME, COL_EMAIL, COL_SUBJECT, COL_BODY = 1, 2, 3, 4
COL_ATT1, COL_ATT2, COL_STATUS            = 5, 6, 7


# ------------------------------- Helpers --------------------------------

def get_exe_dir() -> str:
    """Folder where the .exe (or the .py) lives."""
    if getattr(sys, "frozen", False):                 # PyInstaller bundle
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_settings_path()   : return os.path.join(get_exe_dir(),     SETTINGS_FILE)
def get_template_path()   : return os.path.join(get_exe_dir(),     TEMPLATE_FILE)
def get_attachments_dir() : return os.path.join(get_exe_dir(),     ATTACHMENTS_DIR)


def load_settings():
    """Return the saved settings dict, or None if missing/incomplete."""
    p = get_settings_path()
    if not os.path.exists(p):
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        if all(k in data for k in ("provider", "email", "password")):
            return data
    except Exception:
        pass
    return None


def save_settings(data):
    with open(get_settings_path(), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


# ----------------------------- Bootstrap --------------------------------

def create_template() -> str:
    """Generate Email_List.xlsx with headers + one harmless example row."""
    path = get_template_path()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Email List"

    headers = ["Name", "Email", "Subject", "Message",
               "Attachment 1", "Attachment 2", "Status"]
    for col, h in enumerate(headers, 1):
        c = sheet.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")

    # Example row — pre-marked Sent so the engine skips it on first run.
    example = [
        "(example — delete this row)",
        "you@example.com",
        "Sample subject",
        "Hi there,\n\nThis is what a message looks like.",
        "flyer.pdf",   # drop this file into the Attachments/ folder
        "",
        "Sent",
    ]
    for col, val in enumerate(example, 1):
        sheet.cell(row=2, column=col, value=val)

    for col_letter, width in zip("ABCDEFG", [18, 30, 30, 50, 25, 25, 12]):
        sheet.column_dimensions[col_letter].width = width

    wb.save(path)
    return path


# --------------------------- Credentials --------------------------------

def prompt_credentials(existing=None) -> dict | None:
    """Pop up Tkinter dialogs to collect provider / email / password."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    default_provider = (existing or {}).get("provider", "gmail")
    default_email    = (existing or {}).get("email", "")

    provider = simpledialog.askstring(
        "Email provider",
        "Which email provider are you using?\n\n"
        "  gmail       — Gmail / Google Workspace\n"
        "  outlook     — Outlook.com / Hotmail / Live\n"
        "  office365   — Microsoft 365",
        initialvalue=default_provider, parent=root)
    if not provider:
        root.destroy(); return None
    provider = provider.strip().lower()
    if provider not in PROVIDERS:
        provider = "gmail"

    email = simpledialog.askstring(
        "Email address",
        f"Enter your {provider} email address:",
        initialvalue=default_email, parent=root)
    if not email:
        root.destroy(); return None
    email = email.strip()

    password = simpledialog.askstring(
        "Password",
        f"Enter the password (or App Password) for {email}.\n\n"
        "If you have 2-Step Verification, you'll usually need an App Password:\n"
        "  • Gmail:    https://myaccount.google.com/apppasswords\n"
        "  • Outlook:  account.microsoft.com → Security → App passwords",
        show="*", parent=root)
    root.destroy()

    if not password:
        return None
    return {"provider": provider, "email": email, "password": password}


def _smtp_handshake(settings, timeout=30):
    cfg = PROVIDERS[settings["provider"]]
    server = smtplib.SMTP(cfg["server"], cfg["port"], timeout=timeout)
    server.ehlo()
    server.starttls(context=ssl.create_default_context())
    server.ehlo()
    server.login(settings["email"], settings["password"])
    return server


def test_smtp_connection(settings):
    try:
        s = _smtp_handshake(settings, timeout=15)
        s.quit()
        return True, ""
    except Exception as e:
        return False, str(e)


# ----------------------------- Mail body --------------------------------

def resolve_attachment(name):
    """Find an attachment file by name inside the Attachments/ folder;
    fall back to treating the entry as an absolute path."""
    if not name:
        return None
    s = str(name).strip()
    if not s:
        return None
    candidate = os.path.join(get_attachments_dir(), os.path.basename(s))
    if os.path.exists(candidate):
        return candidate
    if os.path.exists(s):
        return s
    return None


def attach_file(msg, filepath):
    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{os.path.basename(filepath)}"')
    msg.attach(part)


# ------------------------------ Main run --------------------------------

def run_tool():
    print("=" * 44)
    print("   MASS EMAIL MARKETING TOOL  (v4)")
    print("=" * 44)

    exe_dir = get_exe_dir()
    print(f"Working directory: {exe_dir}\n")

    # 1. Make sure template + attachments folder exist.
    template_path   = get_template_path()
    attachments_dir = get_attachments_dir()
    os.makedirs(attachments_dir, exist_ok=True)

    if not os.path.exists(template_path):
        print("No Email_List.xlsx found. Generating a fresh template...")
        create_template()
        print(f"  ✓ Created: {template_path}")
        print(f"  ✓ Folder:  {attachments_dir}\n")
        print("Next steps:")
        print(f"  1. Open  {template_path}")
        print("  2. Fill in one row per recipient (Status = 'Sent' is skipped).")
        print(f"  3. Drop any attachment files into  {attachments_dir}")
        print("  4. Run this tool again.\n")

        # Open the template + attachments folder for the user (Windows).
        try:
            if sys.platform == "win32":
                os.startfile(template_path)
                os.startfile(attachments_dir)
        except Exception:
            pass

        input("Press Enter to close this window...")
        return

    # 2. Load (or prompt for) SMTP credentials.
    settings = load_settings()
    if settings:
        print(f"Saved credentials: {settings['email']} ({settings['provider']})")
        ans = input("Use them? (Y/n): ").strip().lower()
        if ans not in ("", "y", "yes"):
            settings = None

    if not settings:
        creds = prompt_credentials()
        if not creds:
            print("No credentials provided. Exiting.")
            input("Press Enter to close this window...")
            return
        settings = creds
        save_settings(settings)
        print(f"Credentials saved to: {get_settings_path()}")
        print("(Delete that file to reset.)\n")

    # 3. Quick connectivity check before we touch the spreadsheet.
    print(f"Testing {settings['provider']} SMTP...", end=" ")
    ok, err = test_smtp_connection(settings)
    if not ok:
        print(f"FAILED.\n  Reason: {err}\n")
        print("Tip: with 2-Step Verification you usually need an App Password "
              "rather than your account password.")
        input("Press Enter to close this window...")
        return
    print("OK\n")

    # 4. Open the spreadsheet.
    try:
        wb    = openpyxl.load_workbook(template_path)
        sheet = wb.active
    except Exception as e:
        print(f"CRITICAL: Could not open {template_path}\n  Reason: {e}")
        input("Press Enter to close this window...")
        return

    # 5. Connect SMTP for the real send loop.
    try:
        server = _smtp_handshake(settings)
    except Exception as e:
        print(f"CRITICAL: SMTP connection failed.\n  Reason: {e}")
        input("Press Enter to close this window...")
        return

    sent_count = error_count = 0

    try:
        for row in range(2, sheet.max_row + 1):
            name    = sheet.cell(row=row, column=COL_NAME).value
            email   = sheet.cell(row=row, column=COL_EMAIL).value
            subject = sheet.cell(row=row, column=COL_SUBJECT).value
            body    = sheet.cell(row=row, column=COL_BODY).value
            att1    = sheet.cell(row=row, column=COL_ATT1).value
            att2    = sheet.cell(row=row, column=COL_ATT2).value
            status  = sheet.cell(row=row, column=COL_STATUS).value

            # Skip empty rows and rows that were already sent.
            if not email or str(status).strip().lower() == "sent":
                continue

            print(f"  → {email} ... ", end="")

            try:
                msg = MIMEMultipart()
                msg["From"]    = settings["email"]
                msg["To"]      = str(email)
                msg["Subject"] = str(subject) if subject else "Marketing Update"

                text = (f"Dear {name if name else 'Client'},\n\n"
                        f"{body if body else ''}\n\n"
                        "Best regards,\nMarketing Team")
                msg.attach(MIMEText(text, "plain"))

                for att in (att1, att2):
                    p = resolve_attachment(att)
                    if p:
                        attach_file(msg, p)
                    elif att and str(att).strip():
                        print(f"(missing {att}) ", end="")

                server.sendmail(settings["email"], str(email), msg.as_string())

                sheet.cell(row=row, column=COL_STATUS, value="Sent")
                sent_count += 1
                print("OK")
            except Exception as e:
                sheet.cell(row=row, column=COL_STATUS, value="Error")
                error_count += 1
                print(f"FAILED ({e})")
    finally:
        try: server.quit()
        except Exception: pass
        try: wb.save(template_path)
        except Exception as e:
            print(f"\nWarning: could not save Excel — close it next time. ({e})")

    print("\n" + "=" * 44)
    print(f"  DONE.  Sent: {sent_count}   Errors: {error_count}")
    print("=" * 44)


if __name__ == "__main__":
    try:
        run_tool()
    except Exception as e:
        print(f"\nUnexpected error: {e}")
    input("\nPress Enter to close this window...")
